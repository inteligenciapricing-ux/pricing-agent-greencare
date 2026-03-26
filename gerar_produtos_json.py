import json
import re
from pathlib import Path
from openpyxl import load_workbook

ARQUIVO_EXCEL = "data/sku_master.csv.xlsx"
ARQUIVO_JSON = "data/produtos.json"


def limpar_valor(valor):
    if valor is None:
        return ""

    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).replace(".", ",")

    return str(valor).strip()


def normalizar_espacos(texto):
    texto = str(texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def carregar_planilha(caminho_excel):
    caminho = Path(caminho_excel)

    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_excel}")

    wb = load_workbook(caminho_excel, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))

    if not linhas:
        raise ValueError("A planilha está vazia.")

    cabecalho = [limpar_valor(c) for c in linhas[0]]

    colunas_esperadas = [
        "grupo_comparacao",
        "empresa",
        "marca",
        "ean",
        "volume_ml",
        "concentracao_mg_ml",
        "total_mg",
        "tipo_produto",
        "palavras_busca"
    ]

    colunas_faltantes = [c for c in colunas_esperadas if c not in cabecalho]

    if colunas_faltantes:
        raise ValueError(
            f"Colunas faltantes na planilha: {', '.join(colunas_faltantes)}"
        )

    registros = []

    for linha in linhas[1:]:
        if linha is None:
            continue

        registro = {}
        for i, coluna in enumerate(cabecalho):
            valor = linha[i] if i < len(linha) else None
            registro[coluna] = limpar_valor(valor)

        registros.append(registro)

    return registros


def gerar_nome_base(row):
    grupo = normalizar_espacos(row.get("grupo_comparacao", "")).lower()
    tipo_produto = normalizar_espacos(row.get("tipo_produto", "")).upper()

    if "canabidiol" in grupo or "CBD_ISOLADO" in tipo_produto:
        return "Canabidiol"

    if "extrato" in grupo or "cannabis" in grupo:
        return "Extrato de Cannabis Sativa"

    base = row.get("palavras_busca", "")
    base = normalizar_espacos(base)

    if not base:
        return row.get("grupo_comparacao", "").replace("_", " ").strip()

    return base


def gerar_produto_busca(row):
    nome_base = gerar_nome_base(row)
    marca = normalizar_espacos(row.get("marca", ""))
    concentracao = normalizar_espacos(row.get("concentracao_mg_ml", ""))
    volume_ml = normalizar_espacos(row.get("volume_ml", ""))

    partes = []

    if nome_base:
        partes.append(nome_base)

    if concentracao:
        partes.append(f"{concentracao}mg/ml")

    if volume_ml:
        partes.append(f"{volume_ml}ml")

    if marca:
        partes.append(marca)

    texto = " ".join([p for p in partes if p])
    texto = normalizar_espacos(texto)

    return texto


def montar_produtos(registros):
    produtos = []

    for row in registros:
        produto_busca = gerar_produto_busca(row)

        produto = {
            "grupo_comparacao": row.get("grupo_comparacao", ""),
            "empresa": row.get("empresa", ""),
            "marca": row.get("marca", ""),
            "ean": row.get("ean", ""),
            "volume_ml": row.get("volume_ml", ""),
            "concentracao_mg_ml": row.get("concentracao_mg_ml", ""),
            "total_mg": row.get("total_mg", ""),
            "tipo_produto": row.get("tipo_produto", ""),
            "produto_busca": produto_busca,
            "palavras_busca_raw": row.get("palavras_busca", "")
        }

        if produto["grupo_comparacao"] and produto["marca"]:
            produtos.append(produto)

    return produtos


def salvar_json(produtos, caminho_json):
    with open(caminho_json, "w", encoding="utf-8") as f:
        json.dump(produtos, f, ensure_ascii=False, indent=2)


def main():
    print("=" * 70)
    print("GERANDO PRODUTOS.JSON A PARTIR DA PLANILHA MÃE")
    print("=" * 70)

    registros = carregar_planilha(ARQUIVO_EXCEL)
    print(f"Linhas encontradas na planilha: {len(registros)}")

    produtos = montar_produtos(registros)
    print(f"SKUs válidos para exportação: {len(produtos)}")

    salvar_json(produtos, ARQUIVO_JSON)

    print(f"Arquivo gerado com sucesso: {ARQUIVO_JSON}")
    print()

    if produtos:
        print("Primeiro item gerado:")
        print(json.dumps(produtos[0], ensure_ascii=False, indent=2))

    print()
    print("=" * 70)
    print("FINALIZADO")
    print("=" * 70)


if __name__ == "__main__":
    main()